"""BC Configuration Package Generator.

Generates Excel (.xlsx) configuration packages for importing
Cash Receipt Journal entries into Business Central via RapidStart.

Booking logic for CUST-PAYM:
  - For each line item: Credit the principal customer, apply to our doc no.
  - Balance against the bank account.
  - Negative Bruttobetrag (deductions) become debit entries on the customer.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# BC Gen. Journal Line columns for Cash Receipt Journal import
_COLUMNS = [
    ("Journal Template Name", 15),
    ("Journal Batch Name", 15),
    ("Line No.", 10),
    ("Posting Date", 14),
    ("Document Type", 14),
    ("Document No.", 16),
    ("Account Type", 14),
    ("Account No.", 14),
    ("Description", 40),
    ("Amount", 16),
    ("Currency Code", 12),
    ("Applies-to Doc. Type", 18),
    ("Applies-to Doc. No.", 20),
    ("Bal. Account Type", 16),
    ("Bal. Account No.", 16),
    ("External Document No.", 20),
]


def generate_bc_package(
    pass2_results: dict,
    output_dir: str | Path,
    email_subject: str = "",
) -> Optional[Path]:
    """Generate a BC RapidStart Excel configuration package from CUST-PAYM pass2 results.

    Args:
        pass2_results: The pass2_results dict from Pass2CustPaymProcessor.
        output_dir: Directory to write the Excel file.
        email_subject: Original email subject for reference.

    Returns:
        Path to the generated .xlsx file, or None on failure.
    """
    payment_data = pass2_results.get("payment_data")
    if not payment_data:
        logger.error("No payment_data in pass2_results")
        return None

    case_id = pass2_results.get("cust_paym_case_id", "UNKNOWN")
    customer_no = pass2_results.get("principal_customer_no", "")
    bank_account = pass2_results.get("bank_account_no", "BA999")
    currency = pass2_results.get("currency", "EUR")
    payment_date = payment_data.get("payment_date", "")
    wire_transfer_no = payment_data.get("wire_transfer_no", "")
    total_amount = payment_data.get("total_amount", 0.0)
    line_items = payment_data.get("line_items", [])

    if not line_items:
        logger.error("No line items to generate journal entries")
        return None

    # Convert German date DD.MM.YYYY to ISO YYYY-MM-DD for BC
    posting_date = _german_date_to_iso(payment_date) if payment_date else datetime.utcnow().strftime("%Y-%m-%d")
    document_no = wire_transfer_no or f"PAYM-{case_id}-{posting_date}"
    journal_template = "CASHRCPT"
    journal_batch = case_id

    # ── Build workbook ──
    wb = Workbook()

    # Sheet 1: Gen. Journal Line (the import sheet)
    ws = wb.active
    ws.title = "Gen. Journal Line"

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        bottom=Side(style="thin", color="D9D9D9"),
    )
    amount_fmt = '#,##0.00'
    date_fmt = 'YYYY-MM-DD'

    # Write header row
    for col_idx, (col_name, col_width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Write data rows
    line_no = 10000
    row = 2
    for item in line_items:
        our_doc = item.get("our_doc_no", "")
        gross = item.get("gross_amount", 0.0)
        ext_doc = item.get("lidl_doc_no", "")

        if gross == 0.0:
            continue

        # Description
        desc = f"{case_id} {our_doc}"
        if ext_doc:
            desc += f" (Ref: {ext_doc})"

        # BC amount: flip sign (Lidl positive = payment received = BC credit = negative)
        bc_amount = -gross

        # Applies-to doc type
        applies_to_type = "Invoice" if gross > 0 else "Credit Memo"

        values = [
            journal_template,       # Journal Template Name
            journal_batch,          # Journal Batch Name
            line_no,                # Line No.
            posting_date,           # Posting Date
            "Payment",              # Document Type
            document_no,            # Document No.
            "Customer",             # Account Type
            customer_no,            # Account No.
            desc[:100],             # Description
            bc_amount,              # Amount
            currency,               # Currency Code
            applies_to_type,        # Applies-to Doc. Type
            our_doc,                # Applies-to Doc. No.
            "Bank Account",         # Bal. Account Type
            bank_account,           # Bal. Account No.
            ext_doc,                # External Document No.
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = thin_border

        # Format amount column
        ws.cell(row=row, column=10).number_format = amount_fmt

        line_no += 10000
        row += 1

    # Freeze header row
    ws.freeze_panes = "A2"
    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLUMNS))}{row - 1}"

    # Sheet 2: Summary (metadata for reference, not imported)
    ws_summary = wb.create_sheet("Summary")
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 50

    summary_data = [
        ("Case ID", case_id),
        ("Email Subject", email_subject),
        ("Payment Date", posting_date),
        ("Wire Transfer No.", wire_transfer_no),
        ("Total Amount", total_amount),
        ("Currency", currency),
        ("Customer No.", customer_no),
        ("Bank Account", bank_account),
        ("Journal Template", journal_template),
        ("Journal Batch", journal_batch),
        ("Line Items", len(line_items)),
        ("Generated At", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")),
    ]
    for r, (label, value) in enumerate(summary_data, start=1):
        ws_summary.cell(row=r, column=1, value=label).font = Font(bold=True)
        cell = ws_summary.cell(row=r, column=2, value=value)
        if isinstance(value, float):
            cell.number_format = amount_fmt

    # ── Save ──
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    filename = f"BC_CashReceipt_{case_id}_{posting_date}_{wire_transfer_no or 'DRAFT'}.xlsx"
    output_path = output_dir / filename

    wb.save(output_path)
    logger.info(f"BC package generated: {output_path} ({row - 2} journal lines)")
    return output_path


def _german_date_to_iso(date_str: str) -> str:
    """Convert DD.MM.YYYY to YYYY-MM-DD."""
    try:
        dt = datetime.strptime(date_str, "%d.%m.%Y")
        return dt.strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return date_str
