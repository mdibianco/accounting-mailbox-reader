"""Attachment analysis and content extraction."""

import io
import json
import logging
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Optional

from .config import config

logger = logging.getLogger(__name__)


@dataclass
class ExtractedText:
    """Extracted text from an attachment."""

    filename: str
    content_type: str
    text: str
    extraction_method: str
    success: bool
    error: Optional[str] = None


class AttachmentAnalyzer:
    """Analyzes and extracts content from email attachments."""

    SUPPORTED_FORMATS = config.attachment_formats

    def __init__(self):
        """Initialize attachment analyzer."""
        self._load_dependencies()

    def _load_dependencies(self):
        """Attempt to load optional dependencies."""
        self.pypdf = None
        self.pdfplumber = None
        self.openpyxl = None

        try:
            import pypdf

            self.pypdf = pypdf
        except ImportError:
            logger.warning("pypdf not installed, PDF extraction limited")

        try:
            import pdfplumber

            self.pdfplumber = pdfplumber
        except ImportError:
            logger.warning("pdfplumber not installed, PDF extraction limited")

        try:
            import openpyxl

            self.openpyxl = openpyxl
        except ImportError:
            logger.warning("openpyxl not installed, Excel extraction unavailable")

    def analyze(
        self, filename: str, content: bytes, content_type: str
    ) -> ExtractedText:
        """
        Analyze an attachment and extract text.

        Args:
            filename: Name of the attachment
            content: Binary content of the attachment
            content_type: MIME type of the attachment

        Returns:
            ExtractedText object with extracted content
        """
        file_ext = Path(filename).suffix.lower()

        # Check if format is supported
        if file_ext not in self.SUPPORTED_FORMATS:
            return ExtractedText(
                filename=filename,
                content_type=content_type,
                text="",
                extraction_method="unsupported",
                success=False,
                error=f"Unsupported format: {file_ext}",
            )

        # Route to appropriate extractor
        if file_ext == ".pdf":
            return self._extract_pdf(filename, content, content_type)
        elif file_ext in [".xlsx", ".xls"]:
            return self._extract_excel(filename, content, content_type)
        elif file_ext == ".csv":
            return self._extract_csv(filename, content, content_type)
        elif file_ext in [".png", ".jpg", ".jpeg"]:
            return self._extract_image(filename, content, content_type)
        else:
            return ExtractedText(
                filename=filename,
                content_type=content_type,
                text="",
                extraction_method="unknown",
                success=False,
                error="Unknown format",
            )

    def _extract_pdf(
        self, filename: str, content: bytes, content_type: str
    ) -> ExtractedText:
        """Extract text from PDF."""
        try:
            if self.pdfplumber:
                return self._extract_pdf_pdfplumber(filename, content, content_type)
            elif self.pypdf:
                return self._extract_pdf_pypdf(filename, content, content_type)
            else:
                return ExtractedText(
                    filename=filename,
                    content_type=content_type,
                    text="",
                    extraction_method="pdf",
                    success=False,
                    error="No PDF extraction library available",
                )
        except Exception as e:
            logger.error(f"PDF extraction failed for {filename}: {e}")
            return ExtractedText(
                filename=filename,
                content_type=content_type,
                text="",
                extraction_method="pdf",
                success=False,
                error=str(e),
            )

    def _extract_pdf_pdfplumber(
        self, filename: str, content: bytes, content_type: str
    ) -> ExtractedText:
        """Extract PDF using pdfplumber (better accuracy)."""
        pdf_file = io.BytesIO(content)
        text_parts = []

        with self.pdfplumber.open(pdf_file) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                text = page.extract_text() or ""
                if text.strip():
                    text_parts.append(f"--- Page {page_num} ---\n{text}")

        full_text = "\n".join(text_parts)
        return ExtractedText(
            filename=filename,
            content_type=content_type,
            text=full_text,
            extraction_method="pdf_pdfplumber",
            success=bool(full_text.strip()),
        )

    def _extract_pdf_pypdf(
        self, filename: str, content: bytes, content_type: str
    ) -> ExtractedText:
        """Extract PDF using pypdf."""
        pdf_file = io.BytesIO(content)
        text_parts = []

        try:
            reader = self.pypdf.PdfReader(pdf_file)
            for page_num, page in enumerate(reader.pages, 1):
                text = page.extract_text() or ""
                if text.strip():
                    text_parts.append(f"--- Page {page_num} ---\n{text}")
        except Exception as e:
            return ExtractedText(
                filename=filename,
                content_type=content_type,
                text="",
                extraction_method="pdf_pypdf",
                success=False,
                error=str(e),
            )

        full_text = "\n".join(text_parts)
        return ExtractedText(
            filename=filename,
            content_type=content_type,
            text=full_text,
            extraction_method="pdf_pypdf",
            success=bool(full_text.strip()),
        )

    def _extract_excel(
        self, filename: str, content: bytes, content_type: str
    ) -> ExtractedText:
        """Extract text from Excel."""
        if not self.openpyxl:
            return ExtractedText(
                filename=filename,
                content_type=content_type,
                text="",
                extraction_method="excel",
                success=False,
                error="openpyxl not installed",
            )

        try:
            excel_file = io.BytesIO(content)
            workbook = self.openpyxl.load_workbook(excel_file)
            text_parts = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts.append(f"--- Sheet: {sheet_name} ---")

                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join(
                        str(cell) if cell is not None else "" for cell in row
                    )
                    if row_text.strip():
                        text_parts.append(row_text)

            full_text = "\n".join(text_parts)
            return ExtractedText(
                filename=filename,
                content_type=content_type,
                text=full_text,
                extraction_method="excel",
                success=bool(full_text.strip()),
            )
        except Exception as e:
            logger.error(f"Excel extraction failed for {filename}: {e}")
            return ExtractedText(
                filename=filename,
                content_type=content_type,
                text="",
                extraction_method="excel",
                success=False,
                error=str(e),
            )

    def _extract_csv(
        self, filename: str, content: bytes, content_type: str
    ) -> ExtractedText:
        """Extract text from CSV (as-is)."""
        try:
            text = content.decode("utf-8", errors="replace")
            return ExtractedText(
                filename=filename,
                content_type=content_type,
                text=text,
                extraction_method="csv",
                success=True,
            )
        except Exception as e:
            logger.error(f"CSV extraction failed for {filename}: {e}")
            return ExtractedText(
                filename=filename,
                content_type=content_type,
                text="",
                extraction_method="csv",
                success=False,
                error=str(e),
            )

    def _extract_image(
        self, filename: str, content: bytes, content_type: str
    ) -> ExtractedText:
        """Extract text from image (limited, would need OCR for full support)."""
        return ExtractedText(
            filename=filename,
            content_type=content_type,
            text="[Image file - OCR not yet implemented]",
            extraction_method="image",
            success=False,
            error="OCR not yet implemented. Set ocr_enabled=true in config for support.",
        )
